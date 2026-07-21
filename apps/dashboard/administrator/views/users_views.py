import logging
import jdatetime
from django.utils import timezone
from django.views.generic import ListView, View
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ===== Local Imports ===== #
from apps.accounts.permissions import IsTokenJtiActive, HasAdminAccessPermission
from apps.accounts.models import Profile, AuthStatusChoices
from ..forms import UserSearchForm, UserEditForm, ProfileEditForm, AddUserForm
from ..services.email_service import resend_auth_email, send_auth_checked_email
from apps.accounts.services import AmootSMSService

User = get_user_model()

# لاگر اختصاصی احراز هویت
logger = logging.getLogger('user_verification')

# ===== لینک‌های پایه بردکرامب ===== #
BREADCRUMB_HOME = {'label': 'داشبورد', 'url': reverse_lazy('dashboard:index:index')}
BREADCRUMB_USERS = {'label': 'مدیریت کاربران', 'url': reverse_lazy('dashboard:users:admin_users_list')}


# ===== تبدیل تاریخ میلادی به شمسی ===== #
def get_jalali_date(date_obj):
    if not date_obj:
        return ''
    jalali = jdatetime.datetime.fromgregorian(datetime=date_obj.replace(tzinfo=None))
    return jalali.strftime('%Y/%m/%d')


# ================================================== #
# ================ USERS LIST VIEW ================ #
# ================================================== #
class AdminUsersListView(LoginRequiredMixin, IsTokenJtiActive, HasAdminAccessPermission, ListView):
    """ لیست کاربران با جستجو، فیلتر و آمار """

    model = User
    template_name = 'dashboard/users/list.html'
    context_object_name = 'users'
    paginate_by = 15

    def get_queryset(self):
        queryset = User.objects.select_related('profile').order_by('-date_joined')
        form = UserSearchForm(self.request.GET)

        if form.is_valid():
            search = form.cleaned_data.get('search')
            role = form.cleaned_data.get('role')
            auth_status = form.cleaned_data.get('auth_status')
            sort_by = form.cleaned_data.get('sort_by')

            if search:
                queryset = queryset.filter(
                    Q(first_name__icontains=search) |
                    Q(last_name__icontains=search) |
                    Q(phone_number__icontains=search) |
                    Q(email__icontains=search) |
                    Q(profile__medical_code__icontains=search)
                )
            if role:
                queryset = queryset.filter(profile__role=role)
            if auth_status:
                queryset = queryset.filter(profile__auth_status=auth_status)
            if sort_by:
                queryset = queryset.order_by(sort_by)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_form'] = UserSearchForm(self.request.GET or None)
        context['breadcrumb'] = [BREADCRUMB_HOME, {'label': 'مدیریت کاربران', 'url': ''}]
        context['stats'] = {
            'total_users': User.objects.count(),
            'premium_users': Profile.objects.filter(role='premium').count(),
            'pending_verification': Profile.objects.filter(
                role='visitor',
                auth_status=AuthStatusChoices.PENDING.value,
                documents__isnull=False
            ).count(),
            'approved_users': Profile.objects.filter(auth_status='APPROVED').count(),
        }
        for user in context['users']:
            user.jalali_date = get_jalali_date(user.date_joined)
        return context


# ================================================== #
# ================ USER DETAIL VIEW ================ #
# ================================================== #
class UserDetailView(LoginRequiredMixin, HasAdminAccessPermission, View):
    """
    صفحه واحد کاربر: مشاهده کامل، ویرایش inline و احراز هویت.
    POST با فیلد action مدیریت می‌شود: edit | approve | reject
    """

    template_name = 'dashboard/users/detail.html'

    def _context(self, user, **extra):
        profile = user.profile
        try:
            documents = profile.documents.all()
        except Exception:
            documents = []
        context = {
            'target_user': user,
            'profile': profile,
            'documents': documents,
            'is_create': False,
            'user_form': UserEditForm(instance=user),
            'profile_form': ProfileEditForm(instance=profile),
            'breadcrumb': [BREADCRUMB_HOME, BREADCRUMB_USERS, {'label': user.full_name, 'url': ''}],
        }
        context.update(extra)
        return context

    def get(self, request, pk):
        user = get_object_or_404(User.objects.select_related('profile'), pk=pk)
        return render(request, self.template_name, self._context(user))

    def post(self, request, pk):
        user = get_object_or_404(User.objects.select_related('profile'), pk=pk)
        action = request.POST.get('action')

        # ===== ویرایش اطلاعات کاربر ===== #
        if action == 'edit':
            return self._handle_edit(request, user)

        # ===== تایید یا رد احراز هویت ===== #
        if action in ('approve', 'reject'):
            return self._handle_verification(request, user, action)

        messages.error(request, 'عملیات نامعتبر است.')
        return redirect('dashboard:users:admin_user_detail', pk=pk)

    # ===== ذخیره ویرایش inline ===== #
    def _handle_edit(self, request, user):
        user_form = UserEditForm(request.POST, instance=user)
        profile_form = ProfileEditForm(request.POST, instance=user.profile)

        if user_form.is_valid() and profile_form.is_valid():
            with transaction.atomic():
                user_form.save()
                profile_form.save()
            messages.success(request, f'اطلاعات کاربر «{user.full_name}» بروزرسانی شد.')
            return redirect('dashboard:users:admin_user_detail', pk=user.pk)

        errors = {**user_form.errors, **profile_form.errors}
        messages.error(request, f'اطلاعات معتبر نیست: {errors}')
        return render(request, self.template_name, self._context(
            user, user_form=user_form, profile_form=profile_form, open_edit=True
        ))

    # ===== مدیریت احراز هویت ===== #
    def _handle_verification(self, request, user, action):
        profile = user.profile
        sms_service = AmootSMSService()
        logger.info(f"Verification | Admin: {request.user.email} | Target: {user.phone_number} | Action: {action}")

        if action == 'approve':
            medical_code = request.POST.get('medical_code', '').strip()
            profile.auth_status = AuthStatusChoices.APPROVED
            profile.role = 'regular'
            profile.rejection_reason = None
            if medical_code:
                profile.medical_code = medical_code
            profile.save()

            try:
                send_auth_checked_email(user)
            except Exception as e:
                logger.error(f"Approval email failed for {user.email}: {e}")
            try:
                sms_service.send_with_pattern(
                    mobile=str(user.phone_number),
                    pattern_code=4312,
                    values=[user.full_name or 'همکار گرامی'],
                )
            except Exception as e:
                logger.error(f"Approval SMS failed for {user.phone_number}: {e}")

            messages.success(request, f'احراز هویت «{user.full_name}» تایید شد.')

        else:
            reason = request.POST.get('rejection_reason', 'دلیل مشخصی ثبت نشده است.')
            profile.auth_status = AuthStatusChoices.REJECTED
            profile.rejection_reason = reason
            profile.save()

            message_text = (
                f"همکار گرامی {user.full_name}، فرایند احراز هویت شما رد شد.\n"
                f"علت: {reason}\nلطفاً نسبت به احراز هویت مجدد اقدام فرمایید. - دکترکد"
            )
            try:
                resend_auth_email(user)
                sms_service.send_message(mobile=user.phone_number, message_text=message_text)
            except Exception as e:
                logger.error(f"Rejection notify failed for {user.email}: {e}")

            messages.success(request, f'احراز هویت «{user.full_name}» رد شد.')

        return redirect('dashboard:users:admin_user_detail', pk=user.pk)


# ================================================== #
# ================ USER CREATE VIEW ================ #
# ================================================== #
class UserCreateView(LoginRequiredMixin, HasAdminAccessPermission, View):
    """ ایجاد کاربر جدید — همان صفحه detail در حالت ایجاد """

    template_name = 'dashboard/users/detail.html'

    def _context(self, form):
        return {
            'form': form,
            'is_create': True,
            'breadcrumb': [BREADCRUMB_HOME, BREADCRUMB_USERS, {'label': 'افزودن کاربر', 'url': ''}],
        }

    def get(self, request):
        return render(request, self.template_name, self._context(AddUserForm()))

    def post(self, request):
        form = AddUserForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'اطلاعات وارد شده معتبر نیست.')
            return render(request, self.template_name, self._context(form))

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    phone_number=form.cleaned_data['phone_number'],
                    email=form.cleaned_data['email'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    password=form.cleaned_data['password'],
                    is_active=True,
                )
                profile = user.profile
                profile.role = form.cleaned_data['role']
                profile.medical_code = form.cleaned_data['medical_code'] or 'DR-CODE'
                profile.auth_status = 'APPROVED'
                profile.save()
            messages.success(request, f'کاربر «{user.full_name}» با موفقیت اضافه شد.')
            return redirect('dashboard:users:admin_user_detail', pk=user.pk)
        except Exception as e:
            messages.error(request, f'خطا در ایجاد کاربر: {e}')
            return render(request, self.template_name, self._context(form))


# ================================================== #
# ================ USER DELETE VIEW ================ #
# ================================================== #
class UserDeleteView(LoginRequiredMixin, HasAdminAccessPermission, View):
    """ حذف کاربر — پشتیبانی از فرم سرورساید و AJAX """

    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        full_name = user.full_name
        user.delete()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'کاربر «{full_name}» حذف شد.'})

        messages.success(request, f'کاربر «{full_name}» حذف شد.')
        return redirect('dashboard:users:admin_users_list')


# ================================================== #
# ========= PENDING VERIFICATION LIST VIEW ========= #
# ================================================== #
class PendingVerificationListView(LoginRequiredMixin, HasAdminAccessPermission, ListView):
    """ لیست کاربران در صف انتظار احراز هویت """

    model = User
    template_name = 'dashboard/users/verify-user.html'
    context_object_name = 'pending_users'
    paginate_by = 12

    def get_queryset(self):
        return User.objects.select_related('profile').filter(
            profile__auth_status=AuthStatusChoices.PENDING.value,
            profile__documents__isnull=False
        ).order_by('date_joined')


# ================================================== #
# ============= EXPORT USERS TO EXCEL ============= #
# ================================================== #
class ExportUsersToExcelView(LoginRequiredMixin, HasAdminAccessPermission, View):
    """ استخراج اطلاعات کاربران به فرمت Excel """

    def get(self, request):
        role = request.GET.get('role', None)
        users = User.objects.select_related('profile').all()
        if role:
            users = users.filter(profile__role=role)

        wb = Workbook()
        ws = wb.active
        ws.title = "کاربران"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        columns = [
            'نام', 'نام خانوادگی', 'ایمیل', 'شماره تماس',
            'کد نظام پزشکی', 'لینک نظام پزشکی', 'وضعیت احراز هویت',
            'نقش', 'کد معرف', 'دلیل رد هویت', 'تاریخ عضویت'
        ]
        for col_num, title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_num, user in enumerate(users, 2):
            jalali_date = ''
            if user.date_joined:
                jalali_date = jdatetime.datetime.fromgregorian(
                    datetime=user.date_joined
                ).strftime('%Y/%m/%d - %H:%M')

            row_data = [
                user.first_name,
                user.last_name,
                user.email,
                user.phone_number,
                getattr(user.profile, 'medical_code', '') or '',
                getattr(user.profile, 'auth_link', '') or '',
                getattr(user.profile, 'get_auth_status_display', lambda: '')(),
                getattr(user.profile, 'get_role_display', lambda: '')(),
                getattr(user.profile, 'referral_code', '') or '',
                getattr(user.profile, 'rejection_reason', '') or '',
                jalali_date,
            ]
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=str(value))

        widths = [15, 20, 25, 15, 20, 25, 15, 15, 15, 25, 20]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        prefix = f'users_{role}' if role else 'all_users'
        filename = f"{prefix}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
